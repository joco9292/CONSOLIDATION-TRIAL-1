import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from fuzzywuzzy import fuzz, process
import re
import glob
from io import BytesIO
import os
from datetime import datetime
import traceback
import hashlib

# Page configuration
st.set_page_config(
    page_title="Financial Consolidation Tool",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)


# Initialize session state
def init_session_state():
    if 'processed' not in st.session_state:
        st.session_state.processed = False
    if 'consolidated_file' not in st.session_state:
        st.session_state.consolidated_file = None
    if 'processing_logs' not in st.session_state:
        st.session_state.processing_logs = []


init_session_state()

# Title and description
st.title("📊 Financial Statement Consolidation Tool")
st.markdown("""
This tool consolidates financial statements from multiple Excel files into a single comprehensive report.
Upload your files, configure the parameters, and download the consolidated output.
""")

# Add file format requirements info
with st.expander("📋 File Format Requirements", expanded=False):
    st.markdown("""
    **Important:** Ensure your Excel files meet these requirements:

    1. **Source Files**: Should follow naming pattern `fs2025Bedford.xlsx`, `fs2025Dundas.xlsx`, etc.
    2. **Template File**: Must contain sheets named "Income", "Income_statement", "Balance sheet", and "Consolidated Balance Sheet"
    3. **Budget File**: Must have a sheet named `Consolidated {Year}` (e.g., "Consolidated 2025")
    4. **Last Year's File**: Must have sheets named "Income_statement" and "Balance sheet"

    **Common Issues:**
    - Ensure files are not password protected
    - Files must be valid Excel format (.xlsx)
    - Sheet names must match exactly (case-sensitive)
    - Files should not be open in Excel while processing
    """)

# Sidebar for file uploads and configuration
with st.sidebar:
    st.header("📁 File Upload")

    # Source files upload
    source_files = st.file_uploader(
        "Upload Source Financial Files (fs2025*.xlsx)",
        type=['xlsx'],
        accept_multiple_files=True,
        help="Upload multiple Excel files following the naming pattern: fs2025Bedford.xlsx, etc."
    )

    # Template file upload
    template_file = st.file_uploader(
        "Upload Template File",
        type=['xlsx'],
        help="Upload the TEMPLATE_FILE.xlsx"
    )

    # Budget file upload
    budget_file = st.file_uploader(
        "Upload Budget File",
        type=['xlsx'],
        help="Upload the consolidated budget file (e.g., CSIT Consolidated Budget 2025.xlsx)"
    )

    # Last year's consolidated file
    last_year_file = st.file_uploader(
        "Upload Last Year's Consolidated File",
        type=['xlsx'],
        help="Upload the previous year's consolidated financial statement"
    )

    st.divider()

    # Configuration parameters
    st.header("⚙️ Configuration")

    file_year = st.number_input("File Year", min_value=2020, max_value=2030, value=2025)

    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    selected_month = st.selectbox("Selected Month", months, index=5)  # Default to June

    fuzzy_threshold = st.slider(
        "Fuzzy Matching Threshold",
        min_value=50,
        max_value=100,
        value=85,
        help="Higher values require closer matches. Recommended: 85-90 for financial data"
    )

# Main area for processing and results
col1, col2 = st.columns([2, 1])

with col1:
    st.header("📋 Processing Status")

    # Check if all required files are uploaded
    files_ready = all([source_files, template_file, budget_file, last_year_file])

    if not files_ready:
        st.info("👈 Please upload all required files in the sidebar to begin processing.")
        missing_files = []
        if not source_files:
            missing_files.append("- Source financial files")
        if not template_file:
            missing_files.append("- Template file")
        if not budget_file:
            missing_files.append("- Budget file")
        if not last_year_file:
            missing_files.append("- Last year's consolidated file")

        if missing_files:
            st.warning("Missing files:\n" + "\n".join(missing_files))
    else:
        st.success(f"✅ All files uploaded. Ready to process {len(source_files)} source files.")

with col2:
    st.header("📊 File Summary")
    if source_files:
        st.metric("Source Files", len(source_files))
        with st.expander("View uploaded files"):
            for file in source_files:
                st.text(f"• {file.name}")

# Filename to header mapping
filename_to_header = {
    f'fs{file_year}Bedford.xlsx': "Bedford",
    f'fs{file_year}Beechgrove.xlsx': "Scarborough",
    f'fs{file_year}Bering.xlsx': "380 Bering",
    f'fs{file_year}Dundas.xlsx': "Dundas",
    f'fs{file_year}Eastern.xlsx': "Eastern",
    f'fs{file_year}Laird.xlsx': "1 Laird",
    f'fs{file_year}Laird33.xlsx': "33 Laird",
    f'fs{file_year}Lakeshore.xlsx': "Lakeshore",
    f'fs{file_year}Weston.xlsx': "207 Weston",
}


# All the processing functions from the original script
def normalize_label(raw_label):
    if not raw_label:
        return ""
    s = str(raw_label).lower().strip()
    s = re.sub(r'[\/,\(\)\—\–\-\:]', ' ', s)
    s = re.sub(r'\s+', ' ', s)
    synonyms = {
        r'\bindustrial rent\b': 'rental income',
        r'\bretail rental income\b': 'rental income',
        r'\boffice rent\b': 'rental income',
        r'\brental revenue\b': 'rental income',
        r'\brental income\b': 'rental income',
        r'\blease - cp rail\b': 'rent',
        r'\bmerchandise revenue\b': 'merchandise income',
        r'\bmerchandise income\b': 'merchandise income',
        r'\binsurance revenue\b': 'insurance income',
        r'\binsurance income\b': 'insurance income',
        r'\btruck rental\b': 'truck rental',
        r'\buhaul\b': 'truck rental',
        r'\bbad debt\b': 'bad debt',
        r'\bbad debts\b': 'bad debt',
        r'\boffice supplies\b': 'office',
        r'\bprofessional fees\b': 'professional',
        r'\bproperty mgt\. fee\b': 'management fee',
        r'\bmanagement fee\b': 'management fee',
        r'\bamortization/finance cost\b': 'amortization',
        r'\bamortization/finance costs\b': 'amortization',
        r'\bamortization\b': 'amortization',
        r'\bdepreciation\b': 'depreciation',
        r'\bmortgage interest\b': 'mortgage interest',
        r'\bmortgage loan interest\b': 'mortgage interest',
        r'\blegal fee\b': 'legal fees',
        r'\blegal fees\b': 'legal fees',
        r'\blease payment\b': 'lease payment',
        r'\bgain on sale of asset\b': 'gain/loss on investment',
        r'\bgain/loss on investment\b': 'gain/loss on investment',
        r'\bother income\b': 'other income',
        r'\bother revenue\b': 'other income',
        r'\butilities\b': 'utilities',
        r'\bmaintenance & repairs\b': 'maintenance',
        r'\bmaintenance\b': 'maintenance',
        r'\binsurance\b': 'insurance',
        r'\badvertising\b': 'advertising',
        r'\bbank charges\b': 'bank charges',
        r'\brealty taxes\b': 'realty taxes',
        r'\brealty tax\b': 'realty taxes',
        r'\bsalaries & benefits\b': 'salaries',
        r'\bsalaries\b': 'salaries',
        r'\btelephone\b': 'telephone',
        r'\boffice\b': 'office',
        r'\blease\s+cp\s+rail\b': 'rent'
    }
    for pat, replacement in synonyms.items():
        s = re.sub(pat, replacement, s)
    return s.strip()

def normalize_filename(filename):
    """Remove number suffixes like (1), (2) from filenames"""
    import re
    # Remove patterns like " (1)", " (2)", etc. from filename
    cleaned = re.sub(r'\s*\(\d+\)(?=\.xlsx)', '', filename)
    return cleaned


def find_income_sheet(wb):
    for name in wb.sheetnames:
        if "income" in name.lower():
            return name
    return None


def find_anchor_rows(ws):
    rev_start = exp_start = inc_start = None
    pattern_revenue = re.compile(r'^\s*revenue\s*$', re.IGNORECASE)
    pattern_expenses = re.compile(r'^\s*expenses\s*$', re.IGNORECASE)
    pattern_income = re.compile(r'^\s*income\s*$', re.IGNORECASE)

    for row in range(1, ws.max_row + 1):
        val = (ws.cell(row=row, column=1).value or "").strip()
        if rev_start is None and pattern_revenue.match(val):
            rev_start = row
        elif exp_start is None and pattern_expenses.match(val):
            exp_start = row
        elif inc_start is None and pattern_income.match(val):
            inc_start = row
        if rev_start and exp_start and inc_start:
            break

    if rev_start is None:
        for r in range(1, ws.max_row + 1):
            cell = ws.cell(row=r, column=1).value or ""
            if normalize_label(cell) == "rental income":
                rev_start = max(r - 1, 1)
                break

    if exp_start is None and rev_start is not None:
        for r in range(rev_start + 1, ws.max_row + 1):
            cell = ws.cell(row=r, column=1).value or ""
            if normalize_label(cell) == "advertising":
                exp_start = max(r - 1, rev_start + 1)
                break

    if inc_start is None and exp_start is not None:
        for r in range(exp_start + 1, ws.max_row + 1):
            cell = ws.cell(row=r, column=1).value or ""
            if normalize_label(cell) == "management fee":
                inc_start = max(r - 1, exp_start + 1)
                break

    return {"rev_start": rev_start, "exp_start": exp_start, "inc_start": inc_start}


def find_ytd_column(ws):
    for header_row in range(1, 6):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col).value or ""
            if isinstance(cell, str) and re.search(r'y\.?t\.?d', cell, re.IGNORECASE):
                return col
    return 8


def find_month_column(ws):
    for header_row in range(1, 6):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col).value or ""
            if isinstance(cell, str) and re.search(r'Month', cell, re.IGNORECASE):
                return col
    return 2


def parse_income_sheet_ytd(ws):
    anchors = find_anchor_rows(ws)
    rev_row = anchors["rev_start"]
    exp_row = anchors["exp_start"]
    inc_row = anchors["inc_start"]

    if rev_row is None or exp_row is None or inc_row is None:
        return {}, {}, {}

    ytd_col = find_ytd_column(ws)
    rev_dict = {}
    exp_dict = {}
    inc_dict = {}

    for r in range(rev_row + 1, exp_row):
        raw_label = ws.cell(row=r, column=1).value
        if not raw_label:
            continue
        txt = str(raw_label)
        if re.search(r'\btotal revenue\b', txt, re.IGNORECASE):
            continue
        key = normalize_label(txt)
        val = ws.cell(row=r, column=ytd_col).value or 0
        rev_dict[key] = rev_dict.get(key, 0) + val

    for r in range(exp_row + 1, inc_row):
        raw_label = ws.cell(row=r, column=1).value
        if not raw_label:
            continue
        txt = str(raw_label)
        if re.search(r'\btotal operating expenses\b', txt, re.IGNORECASE):
            continue
        key = normalize_label(txt)
        val = ws.cell(row=r, column=ytd_col).value or 0
        exp_dict[key] = exp_dict.get(key, 0) + val

    r = inc_row + 1
    while r <= ws.max_row:
        raw_label = ws.cell(row=r, column=1).value
        if not raw_label:
            break
        txt = str(raw_label)
        if re.search(r'\btotal\b', txt, re.IGNORECASE) or re.search(r'\bnet rental income\b', txt, re.IGNORECASE):
            break
        key = normalize_label(txt)
        val = ws.cell(row=r, column=ytd_col).value or 0
        inc_dict[key] = inc_dict.get(key, 0) + val
        r += 1

    return rev_dict, exp_dict, inc_dict


def parse_income_sheet_month(ws):
    """Parse monthly data from income sheet with comprehensive debugging"""
    anchors = find_anchor_rows(ws)
    rev_row = anchors["rev_start"]
    exp_row = anchors["exp_start"]
    inc_row = anchors["inc_start"]

    if rev_row is None or exp_row is None or inc_row is None:
        st.session_state.processing_logs.append(f"[ERROR] Missing anchor rows: rev={rev_row}, exp={exp_row}, inc={inc_row}")
        return {}, {}, {}

    month_col = find_month_column(ws)
    
    # Debug: Log the column being used and anchor positions
    st.session_state.processing_logs.append(f"[DEBUG] Monthly parsing - Column: {month_col}, Revenue rows: {rev_row+1} to {exp_row-1}")
    
    rev_dict = {}
    exp_dict = {}
    inc_dict = {}

    # ===== REVENUE SECTION =====
    st.session_state.processing_logs.append("[DEBUG] === REVENUE SECTION ===")
    
    # Check what's in the first revenue row
    if rev_row + 1 < exp_row:
        first_label = ws.cell(row=rev_row + 1, column=1).value
        first_value = ws.cell(row=rev_row + 1, column=month_col).value
        st.session_state.processing_logs.append(f"[DEBUG] First revenue row ({rev_row + 1}): label='{first_label}', value='{first_value}'")
    
    for r in range(rev_row + 1, exp_row):
        raw_label = ws.cell(row=r, column=1).value
        if not raw_label:
            continue
            
        txt = str(raw_label)
        
        # Skip total rows
        if re.search(r'\btotal revenue\b', txt, re.IGNORECASE):
            st.session_state.processing_logs.append(f"[DEBUG] Skipping total row {r}: '{txt}'")
            continue
            
        key = normalize_label(txt)
        val = ws.cell(row=r, column=month_col).value or 0
        
        # Log first few revenue items for debugging
        if len(rev_dict) < 3:
            st.session_state.processing_logs.append(f"[DEBUG] Revenue row {r}: '{txt}' -> normalized: '{key}' = {val}")
        
        rev_dict[key] = rev_dict.get(key, 0) + val

    st.session_state.processing_logs.append(f"[DEBUG] Total revenue items parsed: {len(rev_dict)}")
    if rev_dict:
        st.session_state.processing_logs.append(f"[DEBUG] Revenue keys: {list(rev_dict.keys())[:5]}...")  # Show first 5

    # ===== EXPENSES SECTION =====
    st.session_state.processing_logs.append("[DEBUG] === EXPENSES SECTION ===")
    
    # Check what's in the first expense row
    if exp_row + 1 < inc_row:
        first_label = ws.cell(row=exp_row + 1, column=1).value
        first_value = ws.cell(row=exp_row + 1, column=month_col).value
        st.session_state.processing_logs.append(f"[DEBUG] First expense row ({exp_row + 1}): label='{first_label}', value='{first_value}'")
    
    for r in range(exp_row + 1, inc_row):
        raw_label = ws.cell(row=r, column=1).value
        if not raw_label:
            continue
            
        txt = str(raw_label)
        
        # Skip total rows
        if re.search(r'\btotal operating expenses\b', txt, re.IGNORECASE):
            st.session_state.processing_logs.append(f"[DEBUG] Skipping total row {r}: '{txt}'")
            continue
            
        key = normalize_label(txt)
        val = ws.cell(row=r, column=month_col).value or 0
        
        # Log first few expense items for debugging
        if len(exp_dict) < 3:
            st.session_state.processing_logs.append(f"[DEBUG] Expense row {r}: '{txt}' -> normalized: '{key}' = {val}")
        
        exp_dict[key] = exp_dict.get(key, 0) + val

    st.session_state.processing_logs.append(f"[DEBUG] Total expense items parsed: {len(exp_dict)}")
    if exp_dict:
        st.session_state.processing_logs.append(f"[DEBUG] Expense keys: {list(exp_dict.keys())[:5]}...")  # Show first 5

    # ===== INCOME SECTION =====
    st.session_state.processing_logs.append("[DEBUG] === INCOME SECTION ===")
    
    # Check what's in the first income row
    if inc_row + 1 <= ws.max_row:
        first_label = ws.cell(row=inc_row + 1, column=1).value
        first_value = ws.cell(row=inc_row + 1, column=month_col).value
        st.session_state.processing_logs.append(f"[DEBUG] First income row ({inc_row + 1}): label='{first_label}', value='{first_value}'")
    
    r = inc_row + 1
    income_count = 0
    while r <= ws.max_row:
        raw_label = ws.cell(row=r, column=1).value
        if not raw_label:
            break
            
        txt = str(raw_label)
        
        # Stop at total or net rental income rows
        if re.search(r'\btotal\b', txt, re.IGNORECASE) or re.search(r'\bnet rental income\b', txt, re.IGNORECASE):
            st.session_state.processing_logs.append(f"[DEBUG] Stopping at row {r}: '{txt}'")
            break
            
        key = normalize_label(txt)
        val = ws.cell(row=r, column=month_col).value or 0
        
        # Log first few income items for debugging
        if income_count < 3:
            st.session_state.processing_logs.append(f"[DEBUG] Income row {r}: '{txt}' -> normalized: '{key}' = {val}")
            income_count += 1
        
        inc_dict[key] = inc_dict.get(key, 0) + val
        r += 1

    st.session_state.processing_logs.append(f"[DEBUG] Total income items parsed: {len(inc_dict)}")
    if inc_dict:
        st.session_state.processing_logs.append(f"[DEBUG] Income keys: {list(inc_dict.keys())}")  # Show all income keys

    # Final summary
    st.session_state.processing_logs.append("[DEBUG] === MONTHLY PARSING COMPLETE ===")
    st.session_state.processing_logs.append(f"[DEBUG] Summary - Revenue: {len(rev_dict)} items, Expenses: {len(exp_dict)} items, Income: {len(inc_dict)} items")

    return rev_dict, exp_dict, inc_dict


def match_and_write(ws, start_row, end_row, src_dict, target_col_idx):
    candidates = list(src_dict.keys())
    threshold = fuzzy_threshold

    for r in range(start_row, end_row + 1):
        raw_target = (ws.cell(row=r, column=1).value or "").strip()
        key_t = normalize_label(raw_target)
        if not key_t or re.search(r'\btotal\b', raw_target, re.IGNORECASE):
            continue

        cell = ws.cell(row=r, column=target_col_idx)

        if key_t in src_dict:
            cell.value = src_dict[key_t]
            continue

        if candidates:
            result = process.extractOne(key_t, candidates, scorer=fuzz.token_sort_ratio)
            if result:
                best_match, score = result
                if score >= threshold:
                    val = src_dict[best_match]
                    cell.value = val

                    ratio = (score - threshold) / (100 - threshold)
                    green = int(255 * ratio)
                    fill_color = f"FF{green:02X}00"
                    cell.fill = PatternFill(start_color=fill_color,
                                            end_color=fill_color,
                                            fill_type="solid")
                else:
                    cell.value = 0
            else:
                cell.value = 0
        else:
            cell.value = 0


# Then modify the process_one_file_ytd function (around line 415):
def process_one_file_ytd(file_bytes, filename, master_ws, header_row=5):
    # Normalize the filename first
    normalized_filename = normalize_filename(filename)
    
    src_wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    inc_name = find_income_sheet(src_wb)
    if not inc_name:
        st.session_state.processing_logs.append(f"WARNING: no 'INCOME' sheet in {filename}. Skipping.")
        return

    src_ws = src_wb[inc_name]
    rev_dict, exp_dict, inc_dict = parse_income_sheet_ytd(src_ws)
    if not rev_dict and not exp_dict and not inc_dict:
        st.session_state.processing_logs.append(f"WARNING: anchors not found in {filename}. Skipping.")
        return

    # Use normalized filename for lookup
    if normalized_filename in filename_to_header:
        site = filename_to_header[normalized_filename].upper()
    else:
        site = normalized_filename.replace(".xlsx", "").replace(f'fs{file_year}', "").upper()
        if site == "BEECHGROVE":
            site = "SCARBOROUGH"

    target_col = None
    for col in range(1, master_ws.max_column + 1):
        raw_hdr = (master_ws.cell(row=header_row, column=col).value or "").strip()
        raw_hdr_upper = raw_hdr.upper()
        hdr_stripped = re.sub(r'^[\d\s]+', '', raw_hdr_upper)

        if raw_hdr_upper == site or hdr_stripped == site:
            target_col = col
            break

    if not target_col:
        st.session_state.processing_logs.append(f"WARNING: header '{site}' not found in master. Skipping.")
        return

    if "WESTON" in site:
        for d in (rev_dict, exp_dict, inc_dict):
            for k in d:
                d[k] = d[k] / 2

    match_and_write(master_ws, 6, 16, rev_dict, target_col)
    match_and_write(master_ws, 19, 33, exp_dict, target_col)
    match_and_write(master_ws, 38, 46, inc_dict, target_col)
    st.session_state.processing_logs.append(f"✅ YTD data written from {site} → column {target_col}.")

def diagnose_template_structure(master_ws_inc):
    """Diagnose the template structure to find why first row is skipped"""
    st.session_state.processing_logs.append("\n[DIAGNOSTIC] === TEMPLATE STRUCTURE ANALYSIS ===")
    
    # Check YTD section (rows 6-16 for revenue)
    st.session_state.processing_logs.append("\n[DIAGNOSTIC] YTD Revenue Section (rows 6-16):")
    for r in range(6, 17):
        value = master_ws_inc.cell(row=r, column=1).value
        if value:
            normalized = normalize_label(value)
            st.session_state.processing_logs.append(f"  Row {r}: '{value}' -> normalized: '{normalized}'")
    
    # Check Monthly section structure
    st.session_state.processing_logs.append("\n[DIAGNOSTIC] Monthly Section Structure:")
    
    # Check what's in rows 53-56 (around monthly revenue start)
    st.session_state.processing_logs.append("\n[DIAGNOSTIC] Rows 53-56 (Monthly Revenue area):")
    for r in range(53, 57):
        col_a = master_ws_inc.cell(row=r, column=1).value
        col_b = master_ws_inc.cell(row=r, column=2).value
        col_c = master_ws_inc.cell(row=r, column=3).value
        st.session_state.processing_logs.append(f"  Row {r}: A='{col_a}', B='{col_b}', C='{col_c}'")
    
    # Check the actual revenue items in monthly section
    st.session_state.processing_logs.append("\n[DIAGNOSTIC] Monthly Revenue Items (rows 55-65):")
    revenue_items = []
    for r in range(55, 66):
        value = master_ws_inc.cell(row=r, column=1).value
        if value:
            normalized = normalize_label(value)
            revenue_items.append((r, value, normalized))
            st.session_state.processing_logs.append(f"  Row {r}: '{value}' -> normalized: '{normalized}'")
    
    # Check if first item is actually at row 56 instead of 55
    if not master_ws_inc.cell(row=55, column=1).value:
        st.session_state.processing_logs.append("\n[DIAGNOSTIC] ⚠️ Row 55 is EMPTY! First item might be at row 56")
    
    # Compare with YTD section to see if labels match
    st.session_state.processing_logs.append("\n[DIAGNOSTIC] Checking label consistency between YTD and Monthly:")
    ytd_first = master_ws_inc.cell(row=6, column=1).value
    monthly_first_55 = master_ws_inc.cell(row=55, column=1).value
    monthly_first_56 = master_ws_inc.cell(row=56, column=1).value
    
    st.session_state.processing_logs.append(f"  YTD first item (row 6): '{ytd_first}'")
    st.session_state.processing_logs.append(f"  Monthly row 55: '{monthly_first_55}'")
    st.session_state.processing_logs.append(f"  Monthly row 56: '{monthly_first_56}'")
    
    # Check for common template issues
    if monthly_first_55 and "revenue" in str(monthly_first_55).lower():
        st.session_state.processing_logs.append("\n[DIAGNOSTIC] ⚠️ Row 55 appears to be a HEADER ('Revenue') not a data row!")
        st.session_state.processing_logs.append("[DIAGNOSTIC] → Solution: Start monthly revenue at row 56 instead of 55")
    
    st.session_state.processing_logs.append("\n[DIAGNOSTIC] === END ANALYSIS ===\n")


# Add this modified process_one_file_month function
def process_one_file_month(file_bytes, filename, master_ws, header_row=5):
    # Run diagnostic on first file only
    if "Bedford" in filename:
        diagnose_template_structure(master_ws)
    
    # Normalize the filename first
    normalized_filename = normalize_filename(filename)
    
    src_wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    inc_name = find_income_sheet(src_wb)
    if not inc_name:
        return

    src_ws = src_wb[inc_name]
    rev_dict, exp_dict, inc_dict = parse_income_sheet_month(src_ws)
    if not rev_dict and not exp_dict and not inc_dict:
        return

    # Use normalized filename for lookup
    if normalized_filename in filename_to_header:
        site = filename_to_header[normalized_filename].upper()
    else:
        site = normalized_filename.replace(".xlsx", "").replace(f'fs{file_year}', "").upper()
        if site == "BEECHGROVE":
            site = "SCARBOROUGH"

    target_col = None
    for col in range(1, master_ws.max_column + 1):
        raw_hdr = (master_ws.cell(row=header_row, column=col).value or "").strip()
        raw_hdr_upper = raw_hdr.upper()
        hdr_stripped = re.sub(r'^[\d\s]+', '', raw_hdr_upper)

        if raw_hdr_upper == site or hdr_stripped == site:
            target_col = col
            break

    if not target_col:
        return

    if "WESTON" in site:
        for d in (rev_dict, exp_dict, inc_dict):
            for k in d:
                d[k] = d[k] / 2

    # Check if we should adjust the row ranges
    # If row 55 is empty or a header, start from row 56
    first_revenue_row = 53
    first_value = master_ws.cell(row=53, column=1).value
    if not first_value or "revenue" in str(first_value).lower():
        first_revenue_row = 54
        st.session_state.processing_logs.append(f"[DEBUG] Adjusting monthly revenue to start at row {first_revenue_row}")
    
    # Adjusted row ranges
    revenue_end = 65 if first_revenue_row == 53 else 66
    
    match_and_write(master_ws, first_revenue_row, revenue_end, rev_dict, target_col)
    match_and_write(master_ws, 67, 81, exp_dict, target_col)
    match_and_write(master_ws, 86, 94, inc_dict, target_col)
    
    st.session_state.processing_logs.append(f"✅ Monthly data written from {site} → column {target_col}.")
    
# Balance sheet parsing functions
def parse_section(df, start_label, total_label):
    try:
        # Find start index
        start_matches = df[df[0] == start_label].index
        if len(start_matches) == 0:
            st.session_state.processing_logs.append(f"WARNING: Could not find '{start_label}' in balance sheet")
            return {}
        start_idx = start_matches[0]

        # Find end index
        end_matches = df[df[0].astype(str).str.contains(total_label, na=False, case=False)].index
        if len(end_matches) == 0:
            st.session_state.processing_logs.append(f"WARNING: Could not find '{total_label}' in balance sheet")
            return {}
        end_idx = end_matches[0]

        section = df.loc[start_idx + 1: end_idx - 1].reset_index(drop=True)

        sum_col = None
        for col in section.columns[1:]:
            if section[col].apply(lambda x: isinstance(x, (int, float)) and not pd.isna(x)).any():
                sum_col = col
                break
        if sum_col is None:
            sum_col = 1

        result = {}
        i = 0
        n = len(section)
        while i < n:
            cell_i = section.at[i, 0]
            if not isinstance(cell_i, str):
                i += 1
                continue
            header = cell_i.strip()
            header_upper = header.upper()

            if header_upper.startswith("ACCOUNTS RECEIVABLE") and header_upper != "ACCOUNTS RECEIVABLE":
                row_nums = [
                    section.at[i, c]
                    for c in section.columns[1:]
                    if isinstance(section.at[i, c], (int, float)) and not pd.isna(section.at[i, c])
                ]
                if row_nums:
                    val_list = [v for v in row_nums if v != 0]
                    val = float(val_list[-1]) if val_list else 0.0
                    result.setdefault("ACCOUNTS RECEIVABLE", 0.0)
                    result["ACCOUNTS RECEIVABLE"] += val
                i += 1
                continue

            if header_upper == "ACCOUNTS RECEIVABLE":
                result.setdefault("ACCOUNTS RECEIVABLE", 0.0)
                i += 1
                continue

            row_nums = [
                section.at[i, c]
                for c in section.columns[1:]
                if isinstance(section.at[i, c], (int, float)) and not pd.isna(section.at[i, c])
            ]
            nonzero_nums = [v for v in row_nums if v != 0]
            if nonzero_nums:
                result[header] = float(nonzero_nums[-1])
                i += 1
                continue

            total = 0.0
            j = i + 1
            while j < n:
                cell_j = section.at[j, 0]
                if isinstance(cell_j, str):
                    concat = "".join(str(x) for x in section.loc[j] if pd.notna(x))
                    if not any(ch.isdigit() for ch in concat):
                        break
                val_j = section.at[j, sum_col]
                if isinstance(val_j, (int, float)) and not pd.isna(val_j):
                    total += float(val_j)
                j += 1
            result[header] = total
            i = j

        return result
    except Exception as e:
        st.session_state.processing_logs.append(f"ERROR parsing section {start_label}: {str(e)}")
        return {}


def process_balance_sheet_file(file_bytes):
    try:
        xls = pd.ExcelFile(BytesIO(file_bytes))
        if "BALANCE SHEET" in xls.sheet_names:
            sheet = "BALANCE SHEET"
        elif "BALANCESHEET" in xls.sheet_names:
            sheet = "BALANCESHEET"
        else:
            return None

        df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet, header=None)

        # Check if required sections exist
        if "ASSETS" not in df[0].values:
            st.session_state.processing_logs.append("WARNING: 'ASSETS' section not found in balance sheet")
            return None
        if "LIABILITIES" not in df[0].values:
            st.session_state.processing_logs.append("WARNING: 'LIABILITIES' section not found in balance sheet")
            return None
        if "EQUITY" not in df[0].values:
            st.session_state.processing_logs.append("WARNING: 'EQUITY' section not found in balance sheet")
            return None

        assets = parse_section(df, "ASSETS", "TOTAL ASSETS")
        liabilities = parse_section(df, "LIABILITIES", "TOTAL LIABILITIES")
        equity = parse_section(df, "EQUITY", "TOTAL EQUITY")

        # Apply mappings
        hst_keys = [k for k in liabilities if "HST" in k.upper() and "RECOVERABLE" in k.upper()]
        acct_key = next((k for k in liabilities if "ACCOUNTS PAYABLE" in k.upper()), None)
        if acct_key:
            for hst in hst_keys:
                liabilities[acct_key] = liabilities.get(acct_key, 0.0) + liabilities.pop(hst, 0.0)
        else:
            if hst_keys:
                first_hst = hst_keys[0]
                liabilities["ACCOUNTS PAYABLE AND ACCRUED"] = liabilities.pop(first_hst, 0.0)

        if "LOANS PAYABLE" in liabilities:
            liabilities["Due to CSIT Companies"] = liabilities.pop("LOANS PAYABLE")

        if "HOLDBACK PAYABLE" in liabilities:
            liabilities["Deferred Revenue/Deposits"] = liabilities.pop("HOLDBACK PAYABLE")

        if "LOANS RECEIVABLE" in assets:
            assets["Due From CSIT Companies"] = assets.pop("LOANS RECEIVABLE")

        if "PRIVATE ACCOUNTS" in equity:
            equity["Retained Earnings"] = equity.pop("PRIVATE ACCOUNTS")

        return {"ASSETS": assets, "LIABILITIES": liabilities, "EQUITY": equity}
    except Exception as e:
        st.session_state.processing_logs.append(f"ERROR in balance sheet processing: {str(e)}")
        return None


# Main processing function
def process_all_files():
    try:
        st.session_state.processing_logs = []
        current_stage = "Initializing"

        try:
            # Load template workbook
            current_stage = "Loading template file"
            template_file.seek(0)  # Reset file pointer
            master_wb = load_workbook(filename=BytesIO(template_file.read()), data_only=False)

            # Check for required sheets
            required_sheets_template = ["Income", "Income_statement", "Balance sheet", "Balance_sheet"]
            missing_sheets = [sheet for sheet in required_sheets_template if sheet not in master_wb.sheetnames]
            if missing_sheets:
                raise Exception(f"Template file is missing required sheets: {', '.join(missing_sheets)}")

            master_ws_inc = master_wb["Income"]
            master_ws_bs = master_wb["Balance_sheet"]
        except Exception as e:
            raise Exception(f"Error loading template file: {str(e)}")

        # Process Income sheets
        progress_bar = st.progress(0, text="Processing income statements...")
        total_files = len(source_files)

        for idx, file in enumerate(source_files):
            try:
                current_stage = f"Processing income statements from {file.name}"
                file.seek(0)  # Ensure file pointer is at the beginning
                file_bytes = file.read()

                # Process YTD and Monthly data
                process_one_file_ytd(file_bytes, file.name, master_ws_inc)
                process_one_file_month(file_bytes, file.name, master_ws_inc)

                progress = (idx + 1) / (total_files * 2)
                progress_bar.progress(progress, text=f"Processing income statements... {idx + 1}/{total_files}")
            except Exception as e:
                st.session_state.processing_logs.append(f"ERROR processing {file.name}: {str(e)}")

        # Process Balance sheets
        progress_bar.progress(0.5, text="Processing balance sheets...")
        all_summaries = {}

        for idx, file in enumerate(source_files):
            try:
                current_stage = f"Processing balance sheet from {file.name}"
                file.seek(0)  # Reset file pointer
                file_bytes = file.read()
                filename = file.name
                
                # Normalize the filename
                normalized_filename = normalize_filename(filename)
                
                if normalized_filename in filename_to_header:
                    header_label = filename_to_header[normalized_filename]
                    summary = process_balance_sheet_file(file_bytes)
                    if summary is not None:
                        all_summaries[header_label] = summary
                

                progress = 0.5 + (idx + 1) / (total_files * 2) * 0.2
                progress_bar.progress(progress, text=f"Processing balance sheets... {idx + 1}/{total_files}")
            except Exception as e:
                st.session_state.processing_logs.append(f"ERROR processing balance sheet {file.name}: {str(e)}")

        # Build mappings for balance sheet
        current_stage = "Building balance sheet mappings"
        header_row_bs = 5
        row_texts = []
        for row in range(header_row_bs + 1, master_ws_bs.max_row + 1):
            cell_val = master_ws_bs.cell(row=row, column=1).value
            if isinstance(cell_val, str):
                row_texts.append((row, cell_val.strip()))

        row_by_category = {}
        for summary in all_summaries.values():
            for section in ["ASSETS", "LIABILITIES", "EQUITY"]:
                for category in summary[section]:
                    cat_text = category.strip()
                    choices = [text for (_, text) in row_texts]
                    best_match, score = process.extractOne(cat_text, choices, scorer=fuzz.token_set_ratio)
                    if score >= fuzzy_threshold:
                        matches = [r for (r, text) in row_texts if text == best_match]
                        if matches:
                            row_by_category[category] = matches[0]

        col_texts = []
        for col in range(1, master_ws_bs.max_column + 1):
            val = master_ws_bs.cell(row=header_row_bs, column=col).value
            if isinstance(val, str):
                col_texts.append((col, val.strip()))

        col_by_entity = {}
        for entity_label in all_summaries:
            choices = [text for (_, text) in col_texts]
            best_match, score = process.extractOne(entity_label, choices, scorer=fuzz.token_set_ratio)
            if score >= fuzzy_threshold:
                matches = [c for (c, text) in col_texts if text == best_match]
                if matches:
                    col_by_entity[entity_label] = matches[0]

        # Write values into Consolidated Balance Sheet
        current_stage = "Writing balance sheet values"
        for entity_label, summary in all_summaries.items():
            if entity_label not in col_by_entity:
                st.session_state.processing_logs.append(f"Warning: Column for '{entity_label}' not found. Skipping.")
                continue
            target_col = col_by_entity[entity_label]
            for section in ["ASSETS", "LIABILITIES", "EQUITY"]:
                for category, amount in summary[section].items():
                    if category in row_by_category:
                        row_idx = row_by_category[category]
                        if entity_label == "207 Weston":
                            ws_value = f"=({amount})/2"
                        else:
                            ws_value = amount
                        master_ws_bs.cell(row=row_idx, column=target_col, value=ws_value)

        progress_bar.progress(0.7, text="Processing budget data...")

        try:
            # Process budget data
            current_stage = "Processing budget data"
            budget_file.seek(0)  # Reset file pointer

            # First check what sheets are available
            budget_wb = pd.ExcelFile(budget_file)
            sheet_name = f'Consolidated {file_year}'

            if sheet_name not in budget_wb.sheet_names:
                available_sheets = ', '.join(budget_wb.sheet_names)
                raise Exception(
                    f"Budget file does not contain sheet '{sheet_name}'. Available sheets: {available_sheets}")

            df = pd.read_excel(budget_file, sheet_name=sheet_name, header=None)
        except Exception as e:
            raise Exception(f"Error reading budget file: {str(e)}")

        hdr = df[(df[0] == file_year) & (df[2] == 'Jan')].index[0]
        months = df.iloc[hdr, 2:14].tolist()
        month_to_col = {m: c for m, c in zip(months, range(2, 14))}

        sel_col = month_to_col[selected_month]

        start = df[df[0] == 'Rental Revenue'].index[0]
        end_lbl = 'NET PROFIT/(LOSS)'
        if end_lbl not in df[0].values:
            end_lbl = 'NET RENTAL INCOME (LOSS)'
        end = df[df[0] == end_lbl].index[0] + 1

        total_col = month_to_col[months[-1]] + 1

        results = {}
        for r in range(start, end):
            raw = df.iloc[r, 0]
            if isinstance(raw, str):
                m_val = pd.to_numeric(df.iloc[r, sel_col], errors='coerce')
                y_val = pd.to_numeric(df.iloc[r, 2:sel_col + 1], errors='coerce').sum(skipna=True)
                a_val = pd.to_numeric(df.iloc[r, total_col], errors='coerce')
                results[raw.strip()] = (m_val, y_val, a_val)

        # Combine mortgage and CP Rail lines
        if 'Mortgage /Loan Interest' in results and 'Loan Interest (CSIT to Family Mortgage)' in results:
            m1, y1, a1 = results.pop('Mortgage /Loan Interest')
            m2, y2, a2 = results.pop('Loan Interest (CSIT to Family Mortgage)')
            results['Mortgage /Loan Interest'] = (m1 + m2, y1 + y2, a1 + a2)

        if 'CP Rail Lease (Laird)' in results and 'Rent PUD/CSITPM Head Office' in results:
            m1, y1, a1 = results.pop('CP Rail Lease (Laird)')
            m2, y2, a2 = results.pop('Rent PUD/CSITPM Head Office')
            results['CP Rail Lease (Laird)'] = (m1 + m2, y1 + y2, a1 + a2)

        months_count = months.index(selected_month) + 1
        results['Amortization'] = (26500, 26500 * months_count, 26500 * 12)
        results['Depreciation'] = (121963, 121963 * months_count, 121963 * 12)

        progress_bar.progress(0.8, text="Processing last year's data...")

        try:
            # Process last year's data
            current_stage = "Processing last year's income statement"
            last_year_file.seek(0)  # Reset file pointer

            # Check available sheets
            last_year_wb = pd.ExcelFile(last_year_file)
            if 'Income_statement' not in last_year_wb.sheet_names:
                available_sheets = ', '.join(last_year_wb.sheet_names)
                raise Exception(
                    f"Last year's file does not contain 'Income_statement' sheet. Available sheets: {available_sheets}")

            df_last = pd.read_excel(last_year_file, sheet_name='Income_statement', header=None)
        except Exception as e:
            raise Exception(f"Error reading last year's Income_statement: {str(e)}")

        col0 = df_last[0].dropna().astype(str).tolist()

        start_label, score = process.extractOne("RENTAL INCOME", col0)
        if score < 95:
            raise ValueError(f"Couldn't find 'RENTAL INCOME' (@{score}%)")
        start_last = df_last[df_last[0] == start_label].index[0]

        end_label, score = process.extractOne("NET PROFIT/(LOSS)", col0)
        if score < 95:
            end_label, score = process.extractOne("NET RENTAL INCOME (LOSS)", col0)
            if score < 95:
                raise ValueError(f"Couldn't find end-label (@{score}%)")
        end_last = df_last[df_last[0] == end_label].index[0] + 1

        last_results = {}
        for r in range(start_last, end_last):
            raw = df_last.iat[r, 0]
            if isinstance(raw, str):
                am = pd.to_numeric(df_last.iat[r, 1], errors='coerce')
                ay = pd.to_numeric(df_last.iat[r, 7], errors='coerce')
                last_results[raw.strip()] = (am, ay)

        for k1, k2 in [
            ("Mortgage /Loan Interest", "Loan Interest (CSIT to Family Mortgage)"),
            ("CP Rail Lease (Laird)", "Rent PUD/CSITPM Head Office")
        ]:
            if k1 in last_results and k2 in last_results:
                a1, b1 = last_results.pop(k1)
                a2, b2 = last_results.pop(k2)
                last_results[k1] = (a1 + a2, b1 + b2)

        progress_bar.progress(0.9, text="Writing final data...")

        # Write to Income_statement sheet
        if 'Income_statement' not in master_wb.sheetnames:
            raise Exception("Template file is missing 'Income_statement' sheet")
        ws = master_wb['Income_statement']

        template_rows = {
            cell.value.strip().upper(): cell.row
            for cell in ws['A']
            if isinstance(cell.value, str)
        }

        label_map = {
            'Rental Revenue': 'RENTAL INCOME',
            'Merchandise Revenue': 'MERCHANDISE INCOME',
            'Insurance Revenue': 'INSURANCE INCOME',
            'Truck Rental': 'TRUCK RENTAL',
            'Truck and Labour (Mobile)': None,
            'Bad Debts': 'LESS: BAD DEBTS',
            'Other Income': 'OTHER INCOME',
            'Management fee (Weston)': 'MANAGEMENT FEE',
            'Recoveries (Weston)': 'RECOVERIES',
            'Property Management fee - Bedford,': 'PROPERTY MGT. FEE',
            'CP Rail Lease (Laird)': 'RENT',
            'Loss (Gain) on Sale of Equipment': '(GAIN)/LOSS ON INVESTMENT',
            'Merchandise Purchase': 'MERCHANDISE',
            'Advertising': 'ADVERTISING',
            'Utilities': 'UTILITIES',
            'Insurance': 'INSURANCE',
            'Professional Fees': 'PROFESSIONAL FEES',
            'Office Supplies': 'OFFICE SUPPLIES',
            'Bank Charges': 'BANK CHARGES',
            'Realty Tax': 'REALTY TAXES',
            'Maintenance & Repairs': 'MAINTENANCE & REPAIRS',
            'Salaries & Benefits': 'SALARIES & BENEFITS',
            'Telephone': 'TELEPHONE',
            'Rent': 'RENT',
            'Lakeshore Rent Payment': 'LEASE PAYMENT',
            'Minority Interest': 'MINORITY INTEREST',
            'Legal Fees': 'LEGAL FEES',
            'Mortgage /Loan Interest': 'MORTGAGE/LOAN INTEREST',
            'Amortization': 'AMORTIZATION',
            'Depreciation': 'DEPRECIATION',
            'TOTAL REVENUE': None,
            'TOTAL OPERATING EXPENSES': None,
            'NET OPERATING INCOME (LOSS)': None,
            'NET RENTAL INCOME (LOSS)': None,
            'NET PROFIT/(LOSS)': None
        }

        skip_labels = {
            'TOTAL REVENUE',
            'TOTAL OPERATING EXPENSES',
            'NET OPERATING INCOME (LOSS)',
            'NET RENTAL INCOME (LOSS)'
        }

        COL_BUDGET_MONTH = 3
        COL_BUDGET_YTD = 9
        COL_ANNUAL = 15
        COL_LAST_MONTH = 5
        COL_LAST_YTD = 11

        THRESHOLD = 95
        norm_map = {
            k.strip().upper(): (v.strip().upper() if v else None)
            for k, v in label_map.items()
        }

        # Build last_mapped
        last_mapped = {}
        for raw, (am, ay) in last_results.items():
            key = raw.strip().upper()
            if key in norm_map:
                tmpl = norm_map[key]
            elif key in template_rows:
                tmpl = key
            else:
                match, score = process.extractOne(key, template_rows.keys())
                if score >= THRESHOLD:
                    tmpl = match
                else:
                    continue
            if tmpl and tmpl not in skip_labels:
                last_mapped[tmpl] = (am, ay)

        # Write budgets and actuals
        for raw, (m_val, y_val, a_val) in results.items():
            key = raw.strip().upper()
            if key in norm_map and norm_map[key] is None:
                continue

            if key in norm_map:
                lbl = norm_map[key]
            elif key in template_rows:
                lbl = key
            else:
                match, score = process.extractOne(key, template_rows.keys())
                if score >= THRESHOLD:
                    lbl = match
                else:
                    continue

            if lbl in skip_labels:
                continue

            row = template_rows[lbl]
            ws.cell(row=row, column=COL_BUDGET_MONTH).value = float(m_val)
            ws.cell(row=row, column=COL_BUDGET_YTD).value = float(y_val)
            ws.cell(row=row, column=COL_ANNUAL).value = float(a_val)

            if lbl in last_mapped:
                am, ay = last_mapped[lbl]
                ws.cell(row=row, column=COL_LAST_MONTH).value = float(am)
                ws.cell(row=row, column=COL_LAST_YTD).value = float(ay)

        # Apply highlights
        yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for lbl in ('AMORTIZATION', 'DEPRECIATION'):
            if lbl in template_rows:
                ws.cell(row=template_rows[lbl], column=1).fill = yellow

        # Highlight Intercompany row
        bs_ws = master_wb["Balance_sheet"]
        for r in range(1, bs_ws.max_row + 1):
            if str(bs_ws.cell(row=r, column=1).value).strip().upper() == "INTERCOMPANY":
                for c in range(1, 16):
                    bs_ws.cell(row=r, column=c).fill = yellow
                break

        # Process balance sheet from last year
        try:
            current_stage = "Processing last year's balance sheet"
            last_year_file.seek(0)  # Reset file pointer

            # Check if Balance sheet exists
            last_year_wb = pd.ExcelFile(last_year_file)
            sheet = 'Balance sheet'

            if sheet not in last_year_wb.sheet_names:
                # Try alternative naming
                if 'BALANCE SHEET' in last_year_wb.sheet_names:
                    sheet = 'BALANCE SHEET'
                elif 'BalanceSheet' in last_year_wb.sheet_names:
                    sheet = 'BalanceSheet'
                else:
                    available_sheets = ', '.join(last_year_wb.sheet_names)
                    raise Exception(
                        f"Last year's file does not contain 'Balance sheet'. Available sheets: {available_sheets}")

            df0 = pd.read_excel(last_year_file, sheet_name=sheet, header=None)
        except Exception as e:
            raise Exception(f"Error reading last year's Balance sheet: {str(e)}")

        col3 = pd.to_numeric(
            df0.iloc[:, 2].astype(str)
            .str.replace(r'[^0-9.\-]', '', regex=True),
            errors='coerce'
        )

        first_idx = col3.first_valid_index()
        if first_idx is None:
            raise ValueError("No numeric data in col 3!")

        labels = df0.iloc[first_idx:, 0].astype(str).str.strip()
        values = col3.loc[first_idx:]

        balance_dict = {
            lbl: float(val)
            for lbl, val in zip(labels, values)
            if lbl and pd.notna(val)
        }

        current_stage = "Writing last year's balance sheet values"
        if 'Balance sheet' not in master_wb.sheetnames:
            raise Exception("Template file is missing 'Balance sheet' sheet")
        ws = master_wb['Balance sheet']

        for row in ws.iter_rows(min_row=1, max_col=5):
            hdr = row[0].value
            cell_e = row[4]
            if isinstance(hdr, str) and hdr.strip() in balance_dict:
                cell_e.value = balance_dict[hdr.strip()]

        progress_bar.progress(1.0, text="Saving consolidated file...")

        # Save to BytesIO
        output = BytesIO()
        master_wb.save(output)
        output.seek(0)

        st.session_state.consolidated_file = output
        st.session_state.processed = True

        return True

    except Exception as e:
        error_msg = f"Processing Error at stage '{current_stage}': {str(e)}"
        st.error(error_msg)
        with st.expander("Error Details"):
            st.code(traceback.format_exc())
        return False


# Process button
if files_ready:
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        if st.button("🚀 Process Files", type="primary", use_container_width=True):
            with st.spinner("Processing files..."):
                if process_all_files():
                    st.success("✅ Processing complete!")
                    st.balloons()

# Display results
if st.session_state.processed and st.session_state.consolidated_file:
    st.divider()
    st.header("📥 Download Results")

    col1, col2 = st.columns(2)
    with col1:
        st.download_button(
            label="📊 Download Consolidated Financial Report",
            data=st.session_state.consolidated_file,
            file_name=f"Consolidated_Financial_{file_year}_{selected_month}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )

    with col2:
        st.info(f"Report generated for {selected_month} {file_year}")

    # Show processing logs
    if st.session_state.processing_logs:
        with st.expander("📋 Processing Log"):
            for log in st.session_state.processing_logs:
                st.text(log)

# Footer
st.divider()
st.caption("Financial Statement Consolidation Tool v1.0 | Built with Streamlit")